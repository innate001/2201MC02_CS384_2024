from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import csv

# Read student list
with open('/content/stud_list.txt', 'r') as f:
    student_list = [line.strip() for line in f.readlines()]

# Read dates from python_dates.txt
classes_taken_dates = []
classes_missed_dates = []
exams_dates = []

with open('/content/python_dates.txt', 'r') as f:
    for line in f:
        line = line.strip()
        if line.startswith("classes_taken_dates"):
            classes_taken_dates = line.split('=')[1].strip().strip('[]').replace('"', '').split(', ')
        elif line.startswith("classes_missed_dates"):
            classes_missed_dates = line.split('=')[1].strip().strip('[]').replace('"', '').split(', ')
        elif line.startswith("exams_dates"):
            exams_dates = line.split('=')[1].strip().strip('[]').replace('"', '').split(', ')

# All dates combined for lecture_dates
_dates = classes_taken_dates + classes_missed_dates + exams_dates
lecture_dates = [datetime.strptime(date.strip(), '%d/%m/%Y') for date in _dates]
missed_dates = [datetime.strptime(date.strip(), '%d/%m/%Y') for date in classes_missed_dates]
exam_dates = [datetime.strptime(date.strip(), '%d/%m/%Y') for date in exams_dates]

# Input Attendance CSV
att_data = []
with open('/content/input_attendance.csv', 'r') as csvfile:
    csv_reader = csv.reader(csvfile)
    next(csv_reader)  # Skip header row
    for row in csv_reader:
        if row:  # Check if the row is not empty
            ts = datetime.strptime(row[0], '%d/%m/%Y %H:%M:%S')
            roll = row[1].strip()
            att_data.append((ts, roll))

# Initialize attendance record
attendance_record = {student: {date: 0 for date in lecture_dates} for student in student_list}

# Function to count attendance
def count_attendance(timestamps, lecture_date):
    lecture_start = lecture_date + timedelta(hours=18)
    lecture_end = lecture_date + timedelta(hours=20)
    attendance_count = sum(lecture_start <= ts <= lecture_end for ts in timestamps)
    return attendance_count

# Process attendance data
for student in student_list:
    timestamps = [ts for ts, roll in att_data if roll == student]

    for lecture_date in lecture_dates:
        count = count_attendance(timestamps, lecture_date)
        attendance_record[student][lecture_date] = count

# Add additional columns
total_count_dates = len(_dates)
attendance_summary = []

for student in student_list:
    row_data = list(attendance_record[student].values())
    total_attendance_marked = (row_data.count(2) * 2) + row_data.count(1)  # Retained original logic
    total_attendance_allowed = len(classes_taken_dates) * 2
    proxy = sum(1 for x in row_data if x > 2)  # Retained original logic

    attendance_summary.append([student] + row_data + [total_count_dates, total_attendance_marked, total_attendance_allowed, proxy])

# Write to Excel with formatting
wb = Workbook()
ws = wb.active
ws.title = "Attendance Record"

# Define color fills
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Write header
ws.append(["Roll"] + [date.strftime('%d-%m-%Y') for date in lecture_dates] + ["Total count of dates", "Total Attendance Marked", "Total Attendance Allowed", "Proxy"])

# Write data with conditional formatting
for row in attendance_summary:
    ws.append(row)

    for col_idx, value in enumerate(row[1:-4], start=2):  # Attendance columns only
        cell = ws.cell(row=ws.max_row, column=col_idx)
        if value > 2:
            cell.fill = red_fill
        elif value == 2:
            cell.fill = green_fill
        elif value == 1:
            cell.fill = yellow_fill
        elif value == 0:
            cell.fill = PatternFill(fill_type=None)

# Save workbook
output_file = '/content/output.xlsx'
wb.save(output_file)
